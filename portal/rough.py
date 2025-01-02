# Check if there is a parent (user) with the matching reg_number
            try:
                parent = User.objects.get(reg_number=reg_number)
            except User.DoesNotExist:
                parent = None

            if parent:
                # Create or update student record
                # Check if the student already exists
                StudentDet.objects.update_or_create(
                    doc_title=self,
                    name=name,
                    defaults={'reg_number': reg_number, 'grade': grade}
                )



#############################################

from django.db import models
from django.contrib.auth.models import AbstractUser
import pandas as pd
import openpyxl
from django.contrib.auth import get_user_model  # Use this to get the CustomUser model
from django.conf import settings
# Get the custom user model

# Creating extra field(s) for User in the admin panel
class CustomUser(AbstractUser):
    reg_number = models.CharField(max_length=20, unique=True, blank=True, null=True)    

# Add related_name to avoid conflicts with the default User model
    groups = models.ManyToManyField(
        'auth.Group',
        related_name='customuser_set',  # Custom related_name to avoid clashes
        blank=True,
        help_text='The groups this user belongs to.',
        verbose_name='groups'
    )
    user_permissions = models.ManyToManyField(
        'auth.Permission',
        related_name='customuser_permissions_set',  # Custom related_name to avoid clashes
        blank=True,
        help_text='Specific permissions for this user.',
        verbose_name='user permissions'
    )

    def __str__(self):
        return f"{self.username} ({self.reg_number})"
    
# The document to be uploaded containing diff students & their details(excel file)
class DocTitle(models.Model):

    DOCUMENT_TYPE_CHOICES = [
        ('student_details', 'Student Details'),
        ('fee_payment', 'Fee Payment')
    ]

    title = models.CharField(max_length=255)
    type = models.CharField(max_length=20, choices=DOCUMENT_TYPE_CHOICES, default='student_details')  # New field to select document type --
    document = models.FileField(upload_to='docs/')

    def __str__(self):
        return self.title 

    def save(self, *args, **kwargs):
        super().save(*args, **kwargs)
        if self.document: # Check if there is a document uploaded
            self.process_document()
    class Meta:
        verbose_name_plural = 'Documents'

    def process_document(self):
        if self.type == 'student_details':
            self.read_excel_and_create_students()
        elif self.type == 'fee_payment':
            self.read_excel_and_create_fee_payments()

    def read_excel_and_create_students(self):
        User = get_user_model()  # Get the custom user model instead of the default User
        wb = openpyxl.load_workbook(self.document.path)
        sheet = wb.active
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            name, reg_number, grade = row  # Adjust according to your Excel structure

            try:
                parent = User.objects.get(reg_number=reg_number)
            except User.DoesNotExist:
                parent = None

            if parent:
                # Create or update student record
                # Check if the student already exists
                StudentDet.objects.update_or_create(
                    # doc_title=self,
                    name=name,
                    defaults={'reg_number': reg_number, 'grade': grade}
                )   
        
            # Create or update student record
            # Check if the student already exists
            # StudentDet.objects.update_or_create(
                
            #     name=name,
            #     defaults={'reg_number': reg_number, 'grade': grade}
            # )

    def read_excel_and_create_fee_payments(self):
        wb = openpyxl.load_workbook(self.document.path)
        sheet = wb.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            student_name, reg_number, amount, date_paid, balance = row
            
            FeePayment.objects.update_or_create(
                doc_title=self,
                reg_number=reg_number,
                defaults={'student_name': student_name, 'amount': amount, 'date_paid': date_paid, 'balance': balance,}
            )

# The student details (has the same fields as the students excel file)
class StudentDet(models.Model):
    # doc_title = models.ForeignKey(DocTitle, related_name='students', on_delete=models.CASCADE)
    name = models.CharField(max_length=100)
    reg_number = models.CharField(max_length=100, null=True, blank=True)
    grade = models.CharField(max_length=10)

    class Meta:                 
        verbose_name_plural = 'Student Details'

    def __str__(self):
        return self.name

# Fee payment model
class FeePayment(models.Model):
    doc_title = models.ForeignKey(DocTitle, related_name='fee_payments', on_delete=models.CASCADE)
    student_name = models.CharField(max_length=100,  default='Unknown')
    reg_number = models.CharField(max_length=100)
    amount = models.DecimalField(max_digits=10, decimal_places=2)
    date_paid = models.DateField()
    balance = models.DecimalField(max_digits=10, decimal_places=2, default=0.00)

    def __str__(self):
        return f"Payment for {self.reg_number} on {self.date_paid}"
 
class Message(models.Model):
    sender = models.ForeignKey(settings.AUTH_USER_MODEL, related_name='sent_messages', on_delete=models.CASCADE)
    receiver = models.ForeignKey(settings.AUTH_USER_MODEL, related_name='received_messages', on_delete=models.CASCADE)
    message = models.TextField()
    timestamp = models.DateTimeField(auto_now_add=True)
    read = models.BooleanField(default=False)  # New field to track if a message is read

    def __str__(self):
        return f'{self.sender} to {self.receiver}: {self.message[:50]}'

    class Meta:
        ordering = ['timestamp']  # Messages will be ordered by time



    actions = ['process_fee_document']

    @admin.action(description='Process uploaded fee document')
    def process_fee_document(self, request, queryset):
        for payment in queryset:
            if payment.document:
                try:
                    self._process_document(payment.document.path)
                    self.message_user(request, f"Document {payment.document.name} processed successfully.")
                except Exception as e:
                    self.message_user(request, f"Error processing document {payment.document.name}: {e}", level='error')

    def _process_document(self, document_path):
        # Load the Excel file
        df = pd.read_excel(document_path)

        # Ensure the expected columns exist
        expected_columns = {'student_name', 'reg_number', 'amount', 'date_paid', 'balance'}
        if not expected_columns.issubset(df.columns):
            raise ValueError(f"Excel file must contain the following columns: {expected_columns}")

        # Iterate through the rows and create FeePayment instances
        for _, row in df.iterrows():
            FeePayment.objects.create(
                student_name=row.get('student_name', 'Unknown'),
                reg_number=row['reg_number'],
                amount=row['amount'],
                date_paid=row['date_paid'],
                balance=row.get('balance', 0.00)
            )
















from django.http import HttpResponse
from django.shortcuts import render,redirect
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth import update_session_auth_hash
from .models import Message
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout, update_session_auth_hash
from django.http import JsonResponse
from django.core import serializers
from django.contrib.auth import get_user_model
from .models import *
from main.models import StudentAdmission
from django.http import HttpResponse
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from main.models import StudentAdmission
import io


# get_user_model() fetches the custom user model
User = get_user_model()  # This will now refer to 'portal.CustomUser'

@login_required
def portal(request):
    admin_user = User.objects.get(is_superuser=True)  # Assuming admin is the superuser
    messages = Message.objects.filter(sender=request.user, receiver=admin_user) | \
               Message.objects.filter(sender=admin_user, receiver=request.user)
    messages = messages.order_by('timestamp')

    # Get unread messages count
    unread_count = Message.objects.filter(receiver=request.user,sender=request.user, read=False).count()
    
    # Mark unread messages as read when the admin opens the chat
    Message.objects.filter(sender=request.user, receiver=request.user, read=False).update(read=True)

     # Get the current user's reg_number
    reg_number = request.user.reg_number

    # Filter the students based on the reg_number
    students = StudentDet.objects.filter(reg_number=reg_number)
    fee_payments = Fee.objects.filter(reg_number=reg_number)  # Query fee payment data

    if request.method == 'POST':
        message_text = request.POST['message']
        if message_text:
            Message.objects.create(sender=request.user, receiver=admin_user, message=message_text)
            return redirect('portal')  # Reload the page after submitting

    return render(request, 'portal/portal.html', {
        'messages': messages,
        'unread_count': unread_count,  # Pass unread count to the template
        'students': students,
        'fee_payments': fee_payments,
    })

@login_required
def admin(request):
    # Retrieve all student admission records
    admissions = StudentAdmission.objects.all().order_by('-submitted_at')  # Order by submission time, latest first
    return render(request, 'portal/admin.html', {'admissions': admissions})

def download_admissions_pdf(request):
    # Create a byte stream buffer to hold the PDF
    buffer = io.BytesIO()
    pdf = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []

    # Define the header
    styles = getSampleStyleSheet()
    header = Paragraph("Admission Requests Report", styles['Title'])
    subheader = Paragraph("List of all admission requests submitted by prospective students.", styles['BodyText'])
    elements.extend([header, subheader])

    # Define the table data
    admissions = StudentAdmission.objects.all()
    data = [["#", "Name", "Gender", "D.O.B", "Class Applied", "Parent Name", "Contact", "Submitted At"]]
    
    for i, admission in enumerate(admissions, start=1):
        row = [
            i,
            f"{admission.first_name} {admission.last_name}",
            admission.gender,
            admission.date_of_birth.strftime("%Y-%m-%d"),
            admission.applying_class,
            f"{admission.parent_first_name} {admission.parent_last_name}",
            admission.parent_phone,
            admission.submitted_at.strftime("%Y-%m-%d %H:%M:%S"),
        ]
        data.append(row)

    # Create table and set style
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    
    elements.append(table)
    pdf.build(elements)

    # Return the PDF as an HTTP response
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="admission_requests.pdf"'
    return response

@login_required
def delete_chats(request):
    if request.method == 'POST':
        admin_user = User.objects.get(is_superuser=True)
        # Delete all messages between the user and the admin
        Message.objects.filter(sender=request.user, receiver=admin_user).delete()
        Message.objects.filter(sender=admin_user, receiver=request.user).delete()
        return JsonResponse({'status': 'success'})
    
    return JsonResponse({'status': 'error'}, status=403)

def log_in(request):        
    if request.method == 'POST':        
        # email=request.POST.get('email')
        username=request.POST.get('username')
        pass1=request.POST.get('password1')        

        user=authenticate(request, username=username, password=pass1)

        if user is not None:
            login(request,user)
            # Check if the user is a superuser
            if user.is_superuser:
                return redirect('admin')  # Redirect to the admin page if superuser
            else:
                return redirect('portal')  # Redirect to the portal page for other users
        else:
            return redirect('log_in')
        
    return render(request, 'portal/log_in.html')

def logoutUser(request):
    logout(request)
    return redirect('home')

@login_required
def change_password(request):
    if request.method == 'POST':
        form = PasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)  # Keep the user logged in after password change
            messages.success(request, 'Your password has been changed successfully.')
            return redirect('portal')  # Redirect back to the portal after success
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = PasswordChangeForm(request.user)
    
    return render(request, 'portal/change_password.html', {'form': form})

@login_required
def admin_conversations(request):
    if not request.user.is_superuser:
        return redirect('home')

    # Get users who have sent messages to the admin
    users = User.objects.filter(sent_messages__receiver=request.user).distinct()

    users_with_unread = {}
    for user in users:
        unread_count = Message.objects.filter(sender=user, receiver=request.user, read=False).count()
        users_with_unread[user.id] = unread_count

    # If it's an AJAX request, return the list of users and their unread counts
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return JsonResponse({
            'users': list(users.values()),
            'users_with_unread': users_with_unread
        })

    return render(request, 'portal/admin_conversations.html', {
        'users': users,
        'users_with_unread': users_with_unread,
    })

@csrf_exempt
@login_required
def delete_conversation(request):
    if request.method == 'POST' and request.user.is_superuser:
        user_id = request.POST.get('user_id')
        user = User.objects.get(id=user_id)
        # Assuming Message is the model for chat messages
        Message.objects.filter(sender=user, receiver=request.user).delete()
        return JsonResponse({'status': 'success'})

    return JsonResponse({'status': 'error'}, status=403)

def check_new_users(request):
    """Function to check for new users who have sent messages."""
    if request.user.is_superuser:
        new_users = User.objects.filter(sent_messages__receiver=request.user).exclude(id__in=users)
        new_users_list = [{'id': user.id, 'username': user.username} for user in new_users]
        return JsonResponse({'new_users': new_users_list})
    return JsonResponse({'new_users': []})

@login_required
def admin_chat(request, user_id):
    if not request.user.is_superuser:
        return redirect('home')  # Restrict access to admin

    user = User.objects.get(id=user_id)
    admin_user = request.user

    # Fetch all messages between the admin and the user
    messages = Message.objects.filter(sender=user, receiver=admin_user) | \
               Message.objects.filter(sender=admin_user, receiver=user)
    messages = messages.order_by('timestamp')

    # Mark unread messages as read when the admin opens the chat
    Message.objects.filter(sender=user, receiver=admin_user, read=False).update(read=True)

    if request.method == 'POST':
        message_text = request.POST['message']
        if message_text:
            Message.objects.create(sender=admin_user, receiver=user, message=message_text)
            return redirect('admin_chat', user_id=user.id)

    return render(request, 'portal/admin_chat.html', {'messages': messages, 'user': user})

@login_required
def get_messages(request, user_id=None):
    admin_user = request.user if request.user.is_superuser else User.objects.get(is_superuser=True)
    user = User.objects.get(id=user_id) if user_id else request.user

    messages = Message.objects.filter(sender=user, receiver=admin_user) | \
               Message.objects.filter(sender=admin_user, receiver=user)
    messages = messages.order_by('timestamp')

    # Mark all unread messages from admin as read when user opens the chat
    if not request.user.is_superuser:
        Message.objects.filter(sender=admin_user, receiver=user, read=False).update(read=True)

    # Convert the queryset to JSON
    messages_json = serializers.serialize('json', messages)
    return JsonResponse({'messages': messages_json})






results_data = []
    students = Results.objects.values('student').distinct()  # Get distinct students

    for student in students:
        student_results = Results.objects.filter(student=student['student'])
        total_marks = sum(result.marks for result in student_results)
        results_data.append({
            'student': student_results.first().student,
            'results': student_results,
            'total_marks': total_marks
        })








