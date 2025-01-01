from django.db import models
from django.contrib.auth.models import AbstractUser
import pandas as pd
from django.db.models.signals import post_save
from django.dispatch import receiver
import openpyxl
from django.contrib.auth import get_user_model  # Use this to get the CustomUser model
from django.conf import settings
# Get the custom user model

# Creating extra field(s) for User in the admin panel
class CustomUser(AbstractUser):
    reg_number = models.CharField(max_length=20, unique=True, blank=True, null=True)    

# Add related_name to avoid conflicts with the default User model
    groups = models.ManyToManyField(
        'auth.Group',
        related_name='customuser_set',  # Custom related_name to avoid clashes
        blank=True,
        help_text='The groups this user belongs to.',
        verbose_name='groups'
    )
    user_permissions = models.ManyToManyField(
        'auth.Permission',
        related_name='customuser_permissions_set',  # Custom related_name to avoid clashes
        blank=True,
        help_text='Specific permissions for this user.',
        verbose_name='user permissions'
    )

    def __str__(self):
        return f"{self.username} ({self.reg_number})"
    
# The document to be uploaded containing diff students & their details(excel file)
class DocTitle(models.Model):

    DOCUMENT_TYPE_CHOICES = [
        ('student_details', 'Student Details'),
        ('fee_payment', 'Fee Payment')
    ]

    title = models.CharField(max_length=255)
    type = models.CharField(max_length=20, choices=DOCUMENT_TYPE_CHOICES, default='student_details')  # New field to select document type --
    document = models.FileField(upload_to='docs/')

    def __str__(self):
        return self.title 

    def save(self, *args, **kwargs):
        super().save(*args, **kwargs)
        if self.document: # Check if there is a document uploaded
            self.process_document()
    class Meta:
        verbose_name_plural = 'Documents'

    def process_document(self):
        if self.type == 'student_details':
            self.read_excel_and_create_students()
        elif self.type == 'fee_payment':
            self.read_excel_and_create_fee_payments()

    def read_excel_and_create_students(self):
        User = get_user_model()  # Get the custom user model instead of the default User
        wb = openpyxl.load_workbook(self.document.path)
        sheet = wb.active
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            name, reg_number, grade = row  # Adjust according to your Excel structure

            try:
                parent = User.objects.get(reg_number=reg_number)
            except User.DoesNotExist:
                parent = None

            if parent:
                # Create or update student record
                # Check if the student already exists
                StudentDet.objects.update_or_create(
                    # doc_title=self,
                    name=name,
                    defaults={'reg_number': reg_number, 'grade': grade}
                )   
        
            # Create or update student record
            # Check if the student already exists
            # StudentDet.objects.update_or_create(
                
            #     name=name,
            #     defaults={'reg_number': reg_number, 'grade': grade}
            # )

    def read_excel_and_create_fee_payments(self):
        wb = openpyxl.load_workbook(self.document.path)
        sheet = wb.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            student_name, reg_number, amount, date_paid, balance = row
            
            FeePayment.objects.update_or_create(
                doc_title=self,
                reg_number=reg_number,
                defaults={'student_name': student_name, 'amount': amount, 'date_paid': date_paid, 'balance': balance,}
            )

# The student details (has the same fields as the students excel file)
class StudentDet(models.Model):

    DOCUMENT_TYPE_CHOICES = [
        ('playgroup', 'Playgroup'),
        ('pp1', 'pp1'),
        ('pp2', 'pp2'),
        ('pp3', 'pp3'),
        ('pp4', 'pp4'),
        ('pp5', 'pp5'),
        ('pp6', 'pp6'),
        ('pp7', 'pp7'),
        ('pp8', 'pp8'),
        ('grade1', 'grade 1'),
        ('grade2', 'grade 2'),
        ('grade3', 'grade 3'),
        ('grade4', 'grade 4'),
        ('grade5', 'grade 5'),
        ('grade6', 'grade 6'),
        ('grade7', 'grade 7'),
        ('grade8', 'grade 8'),

    ]

    name = models.CharField(max_length=100)
    reg_number = models.CharField(max_length=100, null=True, blank=True)
    # grade = models.CharField(max_length=10)
    class_name = models.CharField(max_length=20, choices=DOCUMENT_TYPE_CHOICES, default='playgroup')  # New field to select document type --

    class Meta:                 
        verbose_name_plural = 'Student Details'

    def __str__(self):
        return self.reg_number

class Message(models.Model):
    sender = models.ForeignKey(settings.AUTH_USER_MODEL, related_name='sent_messages', on_delete=models.CASCADE)
    receiver = models.ForeignKey(settings.AUTH_USER_MODEL, related_name='received_messages', on_delete=models.CASCADE)
    message = models.TextField()
    timestamp = models.DateTimeField(auto_now_add=True)
    read = models.BooleanField(default=False)  # New field to track if a message is read

    def __str__(self):
        return f'{self.sender} to {self.receiver}: {self.message[:50]}'

    class Meta:
        ordering = ['timestamp']  # Messages will be ordered by time

class Fee(models.Model):
    student_name = models.CharField(max_length=100, blank=True, null=True)
    reg_number = models.CharField(max_length=100, blank=True, null=True)
    amount = models.DecimalField(max_digits=10, decimal_places=2, blank=True, null=True)
    date_paid = models.DateField(blank=True, null=True)
    balance = models.DecimalField(max_digits=10, decimal_places=2, blank=True, null=True)
    document = models.FileField(upload_to='fee_documents/', blank=True, null=True)  # For uploading Excel files

    def __str__(self):
        if self.document:
            return f"Fee payments from document: {self.document.name}"
        else:
            return f"Payment for {self.reg_number} on {self.date_paid}"    

    class Meta:
        verbose_name_plural = 'Fees'

    def save(self, *args, **kwargs):
        # Check for skip flag to prevent recursion during bulk creation
        skip_processing = kwargs.pop('skip_processing', False)
        super().save(*args, **kwargs)

        # Process the document only if it's a new instance and not skipped
        if self.pk and self.document and not skip_processing:
            self.process_excel_file()

    def process_excel_file(self):
        try:
            # Load the Excel file
            file_path = self.document.path
            df = pd.read_excel(file_path)

            # Ensure the file has the expected columns
            expected_columns = {'student_name', 'reg_number', 'amount', 'date_paid', 'balance'}
            if not expected_columns.issubset(df.columns):
                raise ValueError(f"The uploaded file must contain these columns: {expected_columns}")

            # Iterate through rows and create new FeePayment instances
            for _, row in df.iterrows():
                fee = Fee(
                    student_name=row.get('student_name', 'Unknown'),
                    reg_number=row['reg_number'],
                    amount=row['amount'],
                    date_paid=row['date_paid'],
                    balance=row.get('balance', 0.00),
                    document=self.document,  # Assign the current document to the new records                    
                )
                fee.save(skip_processing=True)  # Pass the flag only to the `save()` method

            # Check if the current record contains only the document field and delete it if true
            if not self.student_name and not self.reg_number and not self.amount and not self.date_paid and not self.balance:
                self.delete()  # Deletes the current instance

            # Optionally delete the uploaded document after processing
            self.document.delete(save=False)

        except Exception as e:
            raise ValueError(f"Error processing Excel file: {e}")


class HealthProgress(models.Model):
    # Link to a specific student
    student = models.ForeignKey(StudentDet, on_delete=models.CASCADE, related_name="health_records")
    
    # Health status
    health_status = models.CharField(
        max_length=50, 
        choices=[
            ('Excellent', 'Excellent'),
            ('Good', 'Good'),
            ('Average', 'Average'),
            ('Poor', 'Poor')
        ],
        default='Good',
        help_text="Overall health status"
    )
    
    # Additional notes
    notes = models.TextField(blank=True, help_text="Additional health notes")
    
    # Date of the record
    date_recorded = models.DateField(auto_now_add=True)

    class Meta:
        ordering = ['-date_recorded']
        verbose_name = "Health Progress"
        verbose_name_plural = "Health Progress Records"

    def __str__(self):
        return f"{self.student.name}'s Health Record on {self.date_recorded}"



class FeePay(models.Model):
    student = models.ForeignKey(StudentDet, on_delete=models.CASCADE, related_name="fee_records", null=True, blank=True)
    term = models.CharField(
        max_length=50, 
        choices=[
            ('Term 1', 'Term 1'),
            ('Term 2', 'Term 2'),
            ('Term 3', 'Term 3'),            
        ],
        default='Term 1',
    )
    amounts = models.JSONField(        
        default=list,  # Default is an empty list
        help_text="List of payment amounts"
    )
    date_paid = models.DateField(blank=True, null=True) 

    class Meta:
        ordering = ['-date_paid']
        verbose_name = "Fee Payment"
        verbose_name_plural = "Fee Payment Records"

    def __str__(self):
        return f"{self.student.reg_number}'s Fee Payment Record on {self.date_paid}"       

    def total_amount_paid(self):
        """Calculate the total amount paid."""
        return sum(self.amounts)

    def add_payment(self, amount):
        """Add a new payment to the amounts list."""
        self.amounts.append(amount)
        self.save()


class PayFee(models.Model):
    TERM_CHOICES = [
        ('Term 1', 'Term 1'),
        ('Term 2', 'Term 2'),
        ('Term 3', 'Term 3'),
    ]

    student = models.ForeignKey(StudentDet, on_delete=models.CASCADE, related_name="pay_fee_records")
    term = models.CharField(max_length=50, choices=TERM_CHOICES)
    date_paid = models.DateField(blank=True, null=True)
    transaction_mode = models.CharField(max_length=100, blank=True, null=True)
    amount = models.DecimalField(max_digits=10, decimal_places=2, blank=True, null=True)

    class Meta:
        pass

    def __str__(self):
        return f"{self.student.reg_number} - {self.term} ({self.date_paid})"

    

class Results(models.Model):

    TERM_CHOICES = [
        ('Term 1', 'Term 1'),
        ('Term 2', 'Term 2'),
        ('Term 3', 'Term 3'),
    ]       

    TERM_SECTION_CHOICES = [
        ('Opening', 'Opening'),
        ('Mid Term', 'Mid Term'),
        ('End Term', 'End Term'),
    ]

    Subject_Choices = [
        ('Math', 'Mathematics'),
        ('English', 'English'), 
        ('Science', 'Science'),
    ]

    student = models.ForeignKey(StudentDet, on_delete=models.CASCADE, related_name="results_records")
    term = models.CharField(max_length=50, choices=TERM_CHOICES)
    term_section = models.CharField(max_length=50, choices=TERM_SECTION_CHOICES)
    date_recorded = models.DateField(blank=True, null=True)
    subject = models.CharField(max_length=50, choices=Subject_Choices)
    marks =  models.PositiveIntegerField()

    class Meta:
        pass

    def __str__(self):
        return f"{self.student.reg_number} - {self.term} ({self.date_recorded})"
        

class Terms(models.Model):
    term = models.CharField(max_length=50 )
    term_section = models.CharField(max_length=50 )
    start_date = models.DateField(blank=True, null=True)
    end_date = models.DateField(blank=True, null=True)
    year = models.DateField(blank=True, null=True)

    class Meta:
        pass

    def __str__(self):
        return f"{self.term} - {self.term_section} ({self.start_date} to {self.end_date})"
    













